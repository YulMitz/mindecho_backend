"""Export user metrics & scales to Excel.

Reads DATABASE_URL from ../.env. Read-only (SELECT only).

Usage:
    uv run python export_user_metrics.py
    uv run python export_user_metrics.py --user-id <cuid>
    uv run python export_user_metrics.py --out /tmp/foo.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode

import psycopg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

JSON_KEYS = ("physical", "mood", "sleep", "energy", "appetite")


def parse_json_field(raw):
    """Return (value, description). psycopg may already give us a dict."""
    if raw is None:
        return None, None
    if isinstance(raw, dict):
        d = raw
    else:
        try:
            d = json.loads(raw)
        except (TypeError, ValueError):
            return None, str(raw)
    if not isinstance(d, dict):
        return None, str(d)
    return d.get("value"), d.get("description")


def iso(dt):
    if dt is None:
        return None
    if isinstance(dt, datetime):
        return dt.isoformat(sep=" ", timespec="seconds")
    return str(dt)


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append(r)
    # Reasonable column widths
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(str(h)) + 4))


def fetch_metrics(cur, user_id):
    sql = """
        SELECT m."userId", u.email, u.name,
               m.physical, m.mood, m.sleep, m.energy, m.appetite,
               m.entry_date, m.created_at
        FROM mental_health_metrics m
        LEFT JOIN users u ON u."userId" = m."userId"
        {where}
        ORDER BY m."userId", m.entry_date DESC
    """
    params = []
    where = ""
    if user_id:
        where = 'WHERE m."userId" = %s'
        params.append(user_id)
    cur.execute(sql.format(where=where), params)
    rows = []
    for row in cur.fetchall():
        uid, email, name, physical, mood, sleep, energy, appetite, entry_date, created_at = row
        out = [uid, email, name]
        for raw in (physical, mood, sleep, energy, appetite):
            val, desc = parse_json_field(raw)
            out.extend([val, desc])
        out.extend([iso(entry_date), iso(created_at)])
        rows.append(out)
    return rows


def fetch_scale_sessions(cur, user_id):
    sql = """
        SELECT ss.id, ss.user_id, u.email, u.name,
               s.code, s.name, ss.total_score, ss.created_at
        FROM scale_sessions ss
        JOIN scales s ON s.id = ss.scale_id
        LEFT JOIN users u ON u.id = ss.user_id
        {where}
        ORDER BY ss.user_id, ss.created_at DESC
    """
    params = []
    where = ""
    if user_id:
        where = "WHERE ss.user_id = %s"
        params.append(user_id)
    cur.execute(sql.format(where=where), params)
    return [
        [sid, uid, email, name, code, sname, score, iso(created)]
        for sid, uid, email, name, code, sname, score, created in cur.fetchall()
    ]


def fetch_scale_answers(cur, user_id):
    sql = """
        SELECT sa.session_id, sq."order", sq.text, sq.is_reverse, sa.value
        FROM scale_answers sa
        JOIN scale_questions sq ON sq.id = sa.question_id
        JOIN scale_sessions ss ON ss.id = sa.session_id
        {where}
        ORDER BY sa.session_id, sq."order"
    """
    params = []
    where = ""
    if user_id:
        where = "WHERE ss.user_id = %s"
        params.append(user_id)
    cur.execute(sql.format(where=where), params)
    return list(cur.fetchall())


def main():
    ap = argparse.ArgumentParser(description="Export user metrics & scales to Excel.")
    ap.add_argument("--user-id", help="Filter to a single user (User.id cuid)")
    ap.add_argument("--out", help="Output .xlsx path")
    args = ap.parse_args()

    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        sys.exit("ERROR: DATABASE_URL not set in .env")

    # Strip Prisma-specific query params (e.g. `schema=public`) that libpq rejects.
    parts = urlsplit(db_url)
    libpq_allowed = {
        "host", "hostaddr", "port", "dbname", "user", "password", "passfile",
        "channel_binding", "connect_timeout", "client_encoding", "options",
        "application_name", "fallback_application_name", "keepalives",
        "keepalives_idle", "keepalives_interval", "keepalives_count",
        "tcp_user_timeout", "replication", "gssencmode", "sslmode",
        "requiressl", "sslcompression", "sslcert", "sslkey", "sslpassword",
        "sslrootcert", "sslcrl", "sslcrldir", "sslsni", "requirepeer",
        "ssl_min_protocol_version", "ssl_max_protocol_version", "krbsrvname",
        "gsslib", "service", "target_session_attrs", "load_balance_hosts",
    }
    pruned = [(k, v) for k, v in parse_qsl(parts.query) if k in libpq_allowed]
    db_url = urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(pruned), parts.fragment))

    out_path = Path(args.out) if args.out else (
        Path.cwd() / "exports" / f"user_metrics_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Connecting to database (read-only)...")
    with psycopg.connect(db_url) as conn:
        conn.read_only = True
        with conn.cursor() as cur:
            print("Fetching metrics...")
            metrics_rows = fetch_metrics(cur, args.user_id)
            print(f"  {len(metrics_rows)} rows")

            print("Fetching scale sessions...")
            session_rows = fetch_scale_sessions(cur, args.user_id)
            print(f"  {len(session_rows)} rows")

            print("Fetching scale answers...")
            answer_rows = fetch_scale_answers(cur, args.user_id)
            print(f"  {len(answer_rows)} rows")

    wb = Workbook()
    ws = wb.active
    ws.title = "metrics"
    metric_headers = [
        "user_id", "email", "name",
        "physical_value", "physical_desc",
        "mood_value", "mood_desc",
        "sleep_value", "sleep_desc",
        "energy_value", "energy_desc",
        "appetite_value", "appetite_desc",
        "entry_date", "created_at",
    ]
    write_sheet(ws, metric_headers, metrics_rows)

    write_sheet(
        wb.create_sheet("scale_sessions"),
        ["session_id", "user_id", "email", "name",
         "scale_code", "scale_name", "total_score", "created_at"],
        session_rows,
    )

    write_sheet(
        wb.create_sheet("scale_answers"),
        ["session_id", "question_order", "question_text", "is_reverse", "value"],
        answer_rows,
    )

    wb.save(out_path)
    print(f"\n✓ Wrote {out_path}")


if __name__ == "__main__":
    main()
